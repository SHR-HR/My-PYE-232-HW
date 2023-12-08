'''''   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓
Дата выполнения практической работы: 08-09 - ДЕКАБРЯ 2023
'''''
'''
Курс: Разработка Web-приложений на Python, с применением Фреймворка Django
Дисциплина: Основы программирования на Python
'''
'''
Урок от 08.12.2023
Практическая работа №13: Работа с комплексными файлами - excel, json, word. Библиотеки openpyxl, json, docx
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Выполните следующие задания:
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Задание №1

а) Создайте excel файл в операционной системе, заполните его данными в одну строку, от 1 до 10. Во второй строке 
от а и до 10 буквы.
б) Прочитайте эти две строки.
в) Перезапишите их в другой файл, но вертикально, т.е. первая строка – первый столбец.

'''
'''
Решение:
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Вариант - Кода № 1. ↓ 
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
import openpyxl

# Задание а)
# Создаем новый файл
workbook = openpyxl.Workbook()
sheet = workbook.active

# Заполняем первую строку числами от 1 до 10
for i in range(1, 11):
    sheet.cell(row=1, column=i, value=i)

# Заполняем вторую строку буквами от A до J
for i, letter in enumerate('ABCDEFGHIJ', start=1):
    sheet.cell(row=2, column=i, value=letter)

# Сохраняем файл
workbook.save('original_data.xlsx')

# Задание б)
# Загружаем данные из файла
loaded_workbook = openpyxl.load_workbook('original_data.xlsx')
loaded_sheet = loaded_workbook.active

# Читаем данные из первой строки
row_data = [loaded_sheet.cell(row=1, column=i).value for i in range(1, 11)]
print("Первая строка данных:", row_data)

# Читаем данные из второй строки
column_data = [loaded_sheet.cell(row=2, column=i).value for i in range(1, 11)]
print("Вторая строка данных:", column_data)

# Задание в)
# Создаем новый файл для записи вертикальных данных
new_workbook = openpyxl.Workbook()
new_sheet = new_workbook.active

# Записываем данные в новый файл вертикально
for i, value in enumerate(row_data, start=1):
    new_sheet.cell(row=i, column=1, value=value)

# Сохраняем новый файл
new_workbook.save('vertical_data.xlsx')
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Шаг №1: Создание и заполнение файла Excel
'''
# Создаем новый файл
workbook = openpyxl.Workbook()
sheet = workbook.active

# Заполняем первую строку числами от 1 до 10
for i in range(1, 11):
    sheet.cell(row=1, column=i, value=i)

# Заполняем вторую строку буквами от A до J
for i, letter in enumerate('ABCDEFGHIJ', start=1):
    sheet.cell(row=2, column=i, value=letter)

# Сохраняем файл
workbook.save('original_data.xlsx')
'''
Описание:

Создается новый Excel-файл и активный лист (по умолчанию).
В цикле заполняется первая строка числами от 1 до 10, используя метод sheet.cell.
Во второй цикл заполняется вторая строка буквами от A до J.
Файл сохраняется с именем 'original_data.xlsx'.
'''

'''
Шаг №2: Чтение данных из файла Excel
'''
# Загружаем данные из файла
loaded_workbook = openpyxl.load_workbook('original_data.xlsx')
loaded_sheet = loaded_workbook.active

# Читаем данные из первой строки
row_data = [loaded_sheet.cell(row=1, column=i).value for i in range(1, 11)]
print("Первая строка данных:", row_data)

# Читаем данные из второй строки
column_data = [loaded_sheet.cell(row=2, column=i).value for i in range(1, 11)]
print("Вторая строка данных:", column_data)

'''
Описание:

Существующий файл 'original_data.xlsx' загружается с использованием openpyxl.load_workbook.
Читаются данные из первой строки и второй строки с использованием списковых включений.
'''

'''
Шаг №3: Запись данных в новый файл Excel вертикально
'''
# Создаем новый файл для записи вертикальных данных
new_workbook = openpyxl.Workbook()
new_sheet = new_workbook.active

# Записываем данные в новый файл вертикально
for i, value in enumerate(row_data, start=1):
    new_sheet.cell(row=i, column=1, value=value)

# Сохраняем новый файл
new_workbook.save('vertical_data.xlsx')

'''
Описание:

Создается новый Excel-файл и активный лист.
В цикле записываются данные из первой строки вертикально, используя метод new_sheet.cell.
Новый файл сохраняется с именем 'vertical_data.xlsx'.
Таким образом, были выполнены все шаги задания "№1 - а, б, в".
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''↓↓↓
ЕСЛИ НУЖЕН ЧУТОЧКУ ДРУГОЙ РЕЗУЛЬТАТ - ТО: позволил себе небольшое изменение (т.к в оригинальном задании в пункте в) - 
которое звучит так:
"Перезапишите их в другой файл, но вертикально, т.е. первая строка – первый столбец" - я понял это так:
что я беру все данные из ПЕРВОЙ СТРОКИ - (которая у меня записана ГОРИЗОНТАЛЬНО) - ПЕРЕЗАПИСЫВАЮ ЕЕ ЖЕ (эту 1 строчку) в 
НОВЫЙ файл в ПЕРВЫЙ СТОЛБЕЦ - но УЖЕ ВЕРТИКАЛЬНО. А вторую строчку с другими значениями я оставляю только в ОРИГИНАЛЬНОМ
XLSX файле. Если НУЖНО переписать в новый файл ОБЕ строчки - То этот ВАРИАНТ кода ДЛЯ ВАС!=D ↓↓↓↓↓↓↓↓↓↓
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #  НЕМНОГО добавил ИЗМЕНЕНИЙ в одном из ПУНКТОВ - Задания №1  # ~ ~ ~ #
'''

Задание №1

а) Создайте excel файл в операционной системе, заполните его данными в одну строку, от 1 до 10. Во второй строке от а и до 10 буквы.
б) Прочитайте эти две строки.
в) Перезапишите их в другой файл, но вертикально, т.е. первая строка – первый столбец, вторая строка - """второй столбец."""
'''
'''
Вариант - Кода № 2. ↓ 
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
import openpyxl
from openpyxl.utils import get_column_letter

# Шаг 1а: Создание и заполнение файла Excel
def create_and_fill_excel(file_name):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Заполняем первую строку числами от 1 до 10
    sheet.append(list(range(1, 11)))

    # Заполняем вторую строку буквами от A до J
    sheet.append([get_column_letter(i) for i in range(1, 11)])

    # Сохраняем файл
    workbook.save(file_name)

# Шаг 1б: Чтение данных из файла Excel
def read_excel_data(file_name):
    loaded_workbook = openpyxl.load_workbook(file_name)
    loaded_sheet = loaded_workbook.active

    # Читаем данные из первой строки
    row_data = list(loaded_sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]

    # Читаем данные из второй строки
    column_data = list(loaded_sheet.iter_rows(min_row=2, max_row=2, values_only=True))[0]

    return row_data, column_data

# Шаг 1в: Запись данных в новый файл Excel вертикально
def write_data_vertically(output_file, row_data, column_data):
    new_workbook = openpyxl.Workbook()
    new_sheet = new_workbook.active

    # Записываем данные вертикально
    for i, (value1, value2) in enumerate(zip(row_data, column_data), start=1):
        new_sheet.cell(row=i, column=1, value=value1)
        new_sheet.cell(row=i, column=2, value=value2)

    # Сохраняем новый файл
    new_workbook.save(output_file)

# Выполняем шаги
original_file = 'original_data.xlsx'
new_file = 'vertical_data.xlsx'

create_and_fill_excel(original_file)
row_data, column_data = read_excel_data(original_file)
write_data_vertically(new_file, row_data, column_data)
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Подробный "разбор полетов"
'''
'''
Шаг №1 - а: Создание и заполнение файла Excel
'''
'''
Определение: В этом шаге мы создаем новый файл Excel и заполняем его данными в одну 
строку от 1 до 10 и во второй строке от A до J.

Пример написания:
'''
def create_and_fill_excel(file_name):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Заполняем первую строку числами от 1 до 10
    sheet.append(list(range(1, 11)))

    # Заполняем вторую строку буквами от A до J
    sheet.append([get_column_letter(i) for i in range(1, 11)])

    # Сохраняем файл
    workbook.save(file_name)
'''
Подробное описание: 

Мы используем библиотеку openpyxl для создания нового файла Excel. 
Сначала мы создаем экземпляр рабочей книги и активного листа. 
Затем мы добавляем данные в первую строку, используя функцию append, генерируя числа от 1 до 10 с помощью range. 
После этого мы добавляем вторую строку, используя буквы от A до J, преобразованные с помощью функции get_column_letter. 
Наконец, мы сохраняем файл.
'''
'''
Шаг №1 - б: Чтение данных из файла Excel
'''
'''
Определение: В этом шаге мы читаем данные из созданного файла Excel, конкретно данные из первой и второй строк.

Пример написания:
'''
def read_excel_data(file_name):
    loaded_workbook = openpyxl.load_workbook(file_name)
    loaded_sheet = loaded_workbook.active

    # Читаем данные из первой строки
    row_data = list(loaded_sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]

    # Читаем данные из второй строки
    column_data = list(loaded_sheet.iter_rows(min_row=2, max_row=2, values_only=True))[0]

    return row_data, column_data
'''
Подробное описание: Мы используем функцию load_workbook из библиотеки openpyxl, чтобы загрузить созданный файл Excel.
Затем мы используем iter_rows для чтения данных из первой и второй строк с помощью параметров min_row и max_row.
Мы также используем values_only=True, чтобы получить только значения из ячеек. Возвращаемые данные - это списки 
значений из соответствующих строк.
'''
'''
Шаг №1 - в: Запись данных в новый файл Excel вертикально
'''
'''
Определение: В этом шаге мы перезаписываем данные в другой файл Excel, но уже вертикально, 
так чтобы первая строка стала первым столбцом, а вторая строка - вторым столбцом.

Пример написания:
'''
def write_data_vertically(output_file, row_data, column_data):
    new_workbook = openpyxl.Workbook()
    new_sheet = new_workbook.active

    # Записываем данные вертикально
    for i, (value1, value2) in enumerate(zip(row_data, column_data), start=1):
        new_sheet.cell(row=i, column=1, value=value1)
        new_sheet.cell(row=i, column=2, value=value2)

    # Сохраняем новый файл
    new_workbook.save(output_file)
'''
Подробное описание: 

Мы создаем новый файл Excel с помощью openpyxl.Workbook(). 
Затем мы используем цикл и enumerate, чтобы пройти по парам значений из первой и второй строк. 
Мы записываем значения в новый файл, используя cell и указывая соответствующие строки и столбцы. 
Наконец, мы сохраняем новый файл.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Вариант № 2 - Кода № 2. ↓ 


Этот вариант кода также выполняет те же шаги, что и предыдущий, но использует другой подход к чтению 
данных из строк и их записи вертикально. Вместо использования iter_rows, мы просто проходим по 
ячейкам каждой строки и извлекаем их значения.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
import openpyxl
from openpyxl.utils import get_column_letter

# Шаг 1а: Создание и заполнение файла Excel
def create_and_fill_excel(file_name):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Заполняем первую строку числами от 1 до 10
    sheet.append(list(range(1, 11)))

    # Заполняем вторую строку буквами от A до J
    sheet.append([get_column_letter(i) for i in range(1, 11)])

    # Сохраняем файл
    workbook.save(file_name)

# Шаг 1б: Чтение данных из файла Excel
def read_excel_data(file_name):
    loaded_workbook = openpyxl.load_workbook(file_name)
    loaded_sheet = loaded_workbook.active

    # Читаем данные из первой строки
    row_data = [cell.value for cell in loaded_sheet[1]]

    # Читаем данные из второй строки
    column_data = [cell.value for cell in loaded_sheet[2]]

    return row_data, column_data

# Шаг 1в: Запись данных в новый файл Excel вертикально
def write_data_vertically(output_file, row_data, column_data):
    new_workbook = openpyxl.Workbook()
    new_sheet = new_workbook.active

    # Записываем данные вертикально
    for i in range(1, 11):
        new_sheet.cell(row=i, column=1, value=row_data[i-1])
        new_sheet.cell(row=i, column=2, value=column_data[i-1])

    # Сохраняем новый файл
    new_workbook.save(output_file)

# Выполняем шаги
original_file = 'original_data.xlsx'
new_file = 'vertical_data.xlsx'

create_and_fill_excel(original_file)
row_data, column_data = read_excel_data(original_file)
write_data_vertically(new_file, row_data, column_data)
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Шаг 1 - пункт а): Создание и заполнение файла Excel

Определение: 
В этом шаге мы создаем новый файл Excel и заполняем его данными
в одну строку от 1 до 10 и во второй строке от A до J.

Пример написания:
'''
def create_and_fill_excel(file_name):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Заполняем первую строку числами от 1 до 10
    sheet.append(list(range(1, 11)))

    # Заполняем вторую строку буквами от A до J
    sheet.append([get_column_letter(i) for i in range(1, 11)])

    # Сохраняем файл
    workbook.save(file_name)
'''
Подробное описание: 

Мы используем библиотеку openpyxl для создания нового файла Excel. 
Создаем новую рабочую книгу и активный лист. Заполняем первую строку числами от 1 до 10 с помощью range и append. 
Заполняем вторую строку буквами от A до J, используя get_column_letter для генерации букв. Сохраняем созданный файл.
'''
'''
Шаг 1 - пункт б): Чтение данных из файла Excel

Определение: 
В этом шаге мы читаем данные из созданного файла Excel, конкретно данные из первой и второй строк.

Пример написания:
'''
def read_excel_data(file_name):
    loaded_workbook = openpyxl.load_workbook(file_name)
    loaded_sheet = loaded_workbook.active

    # Читаем данные из первой строки
    row_data = [cell.value for cell in loaded_sheet[1]]

    # Читаем данные из второй строки
    column_data = [cell.value for cell in loaded_sheet[2]]

    return row_data, column_data
'''
Подробное описание: 

Мы используем load_workbook из openpyxl для загрузки файла Excel. 
Затем мы проходим по каждой ячейке первой и второй строк с использованием списковых выражений и извлекаем значения 
с помощью cell.value. Возвращаемые данные - это списки значений из соответствующих строк.
'''
'''
Шаг 1 - пункт d): Запись данных в новый файл Excel вертикально

Определение: 
В этом шаге мы перезаписываем данные в другой файл Excel, но уже вертикально, 
так чтобы первая строка стала первым столбцом, а вторая строка - вторым столбцом.

Пример написания:
'''
def write_data_vertically(output_file, row_data, column_data):
    new_workbook = openpyxl.Workbook()
    new_sheet = new_workbook.active

    # Записываем данные вертикально
    for i in range(1, 11):
        new_sheet.cell(row=i, column=1, value=row_data[i-1])
        new_sheet.cell(row=i, column=2, value=column_data[i-1])

    # Сохраняем новый файл
    new_workbook.save(output_file)
'''
Подробное описание: 

Мы создаем новый файл Excel с помощью openpyxl.Workbook(). 
Затем мы используем цикл for и range для прохода по значениям от 1 до 10. 
Мы используем cell для записи данных вертикально в новый файл, указывая соответствующие строки и столбцы. 
Наконец, мы сохраняем новый файл.
'''




