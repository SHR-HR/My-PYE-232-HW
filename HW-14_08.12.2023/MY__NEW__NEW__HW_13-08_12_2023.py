'''''   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓
Дата выполнения домашней работы: 08-09 - ДЕКАБРЯ 2023
'''''
'''
Курс: Разработка Web-приложений на Python, с применением Фреймворка Django
Дисциплина: Основы программирования на Python
'''
'''
Урок от 08.12.2023
Домашняя работа №13: Работа с комплексными файлами - excel, json, word. Библиотеки openpyxl, json, docx
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Выполните следующие задания:
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Задание №1

а) Прочитайте из трёх excel файлов (заранее создайте их, внутри 1111, 2222, 3333).
б) Отсортируйте полученную матрицу в порядке убывания.
в) Запишите это в один файл, изменив шрифт и обернув в границы.
'''
'''
Решение:
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Вариант - Кода № 1. ↓ 

Этот код выполняет все шаги задания в одном блоке, 
читая данные из трех файлов Excel, 
сортируя их в порядке убывания и записывая 
в новый файл с измененным шрифтом и границами.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
import openpyxl
from openpyxl.styles import Font, Border, Side

def read_excel_files(file_names):
    data_matrix = []

    for file_name in file_names:
        loaded_workbook = openpyxl.load_workbook(file_name)
        loaded_sheet = loaded_workbook.active

        # Читаем данные из всех строк файла и добавляем в матрицу
        for row in loaded_sheet.iter_rows(values_only=True):
            data_matrix.append(row)

    return data_matrix

def sort_matrix_descending(matrix):
    # Используем sorted для сортировки по первому элементу в каждой строке в порядке убывания
    sorted_matrix = sorted(matrix, key=lambda x: x[0], reverse=True)
    return sorted_matrix

def write_to_excel(output_file, data_matrix):
    new_workbook = openpyxl.Workbook()
    new_sheet = new_workbook.active

    # Записываем отсортированные данные в новый файл
    for row_num, row_data in enumerate(data_matrix, start=1):
        for col_num, value in enumerate(row_data, start=1):
            new_sheet.cell(row=row_num, column=col_num, value=value)

    # Изменим шрифт и добавим границы
    for row in new_sheet.iter_rows(min_row=1, max_row=len(data_matrix), min_col=1, max_col=len(data_matrix[0])):
        for cell in row:
            cell.font = Font(bold=True)
            cell.border = Border(bottom=Side(style='thin'))

    # Сохраняем новый файл
    new_workbook.save(output_file)

# Пример использования:
files_to_read = ['1.xlsx', '2.xlsx', '3.xlsx']
output_file = 'result.xlsx'

matrix = read_excel_files(files_to_read)
sorted_matrix = sort_matrix_descending(matrix)
write_to_excel(output_file, sorted_matrix)
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Шаг 1: Импорт библиотек
'''
import openpyxl
from openpyxl.styles import Font, Border, Side
'''
В этом шаге мы импортируем необходимые модули из библиотеки openpyxl, 
также импортируем Font, Border и Side для работы со стилями ячеек.
'''
'''
Шаг 2: Чтение данных из трех Excel файлов
'''
def read_excel_files(file_names):
    data_matrix = []

    for file_name in file_names:
        loaded_workbook = openpyxl.load_workbook(file_name)
        loaded_sheet = loaded_workbook.active

        # Читаем данные из всех строк файла и добавляем в матрицу
        for row in loaded_sheet.iter_rows(values_only=True):
            data_matrix.append(row)

    return data_matrix
'''
В этой функции читаем данные из трех Excel файлов. 
Мы используем цикл для перебора имен файлов, загружаем рабочую книгу с помощью openpyxl.load_workbook, 
получаем активный лист и читаем все строки с использованием iter_rows. 
Значения каждой строки добавляются в матрицу data_matrix.
'''
'''
Шаг 3: Сортировка матрицы в порядке убывания
'''
def sort_matrix_descending(matrix):
    # Используем sorted для сортировки по первому элементу в каждой строке в порядке убывания
    sorted_matrix = sorted(matrix, key=lambda x: x[0], reverse=True)
    return sorted_matrix
'''
В этой функции мы используем встроенную функцию sorted для сортировки матрицы в порядке убывания.
Ключ сортировки указывает, что мы сортируем по первому элементу (столбцу) в каждой строке.
'''
'''
Шаг 4: Запись отсортированных данных в новый файл
'''
def write_to_excel(output_file, data_matrix):
    new_workbook = openpyxl.Workbook()
    new_sheet = new_workbook.active

    # Записываем отсортированные данные в новый файл
    for row_num, row_data in enumerate(data_matrix, start=1):
        for col_num, value in enumerate(row_data, start=1):
            new_sheet.cell(row=row_num, column=col_num, value=value)

    # Изменим шрифт и добавим границы
    for row in new_sheet.iter_rows(min_row=1, max_row=len(data_matrix), min_col=1, max_col=len(data_matrix[0])):
        for cell in row:
            cell.font = Font(bold=True)
            cell.border = Border(bottom=Side(style='thin'))

    # Сохраняем новый файл
    new_workbook.save(output_file)
'''
Эта функция создает новую рабочую книгу и активный лист.
Затем записывает отсортированные данные из матрицы в новый файл, используя два вложенных цикла.
После этого изменяет шрифт на жирный (Font(bold=True)) и добавляет границы (Border(bottom=Side(style='thin'))) 
к каждой ячейке. Наконец, сохраняет новый файл.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Вариант № 2 - Кода № 1. ↓ 
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
import openpyxl
import pandas as pd
from openpyxl.styles import Font, Border, Side

def read_excel_files(file_names):
    data_matrix = []

    for file_name in file_names:
        loaded_workbook = openpyxl.load_workbook(file_name)
        loaded_sheet = loaded_workbook.active

        # Читаем данные из всех строк файла и добавляем в матрицу
        for row in loaded_sheet.iter_rows(values_only=True):
            data_matrix.append(row)

    return data_matrix

def sort_matrix_descending(matrix):
    # Используем sorted для сортировки по первому элементу в каждой строке в порядке убывания
    sorted_matrix = sorted(matrix, key=lambda x: x[0], reverse=True)
    return sorted_matrix

def write_to_excel(output_file, data_matrix):
    new_workbook = openpyxl.Workbook()
    new_sheet = new_workbook.active

    # Записываем отсортированные данные в новый файл
    for row_num, row_data in enumerate(data_matrix, start=1):
        for col_num, value in enumerate(row_data, start=1):
            new_sheet.cell(row=row_num, column=col_num, value=value)

    # Изменим шрифт и добавим границы
    for row in new_sheet.iter_rows(min_row=1, max_row=len(data_matrix), min_col=1, max_col=len(data_matrix[0])):
        for cell in row:
            cell.font = Font(bold=True)
            cell.border = Border(left=Side(border_style='thin'),
                                 right=Side(border_style='thin'),
                                 top=Side(border_style='thin'),
                                 bottom=Side(border_style='thin'))

    # Сохраняем новый файл
    new_workbook.save(output_file)

# Пример использования
files_to_read = ['1.xlsx', '2.xlsx', '3.xlsx']
output_file = 'result.xlsx'

matrix = read_excel_files(files_to_read)
sorted_matrix = sort_matrix_descending(matrix)
write_to_excel(output_file, sorted_matrix)

print(f'Данные были успешно записаны в файл: {output_file}')
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Шаг №1: Определение - Чтение данных из трех excel файлов
'''
def read_excel_files(file_names):
    data_matrix = []

    for file_name in file_names:
        loaded_workbook = openpyxl.load_workbook(file_name)
        loaded_sheet = loaded_workbook.active

        # Читаем данные из всех строк файла и добавляем в матрицу
        for row in loaded_sheet.iter_rows(values_only=True):
            data_matrix.append(row)

    return data_matrix
'''
Описание:

Функция read_excel_files принимает список имен файлов file_names.
Создается пустая матрица data_matrix, в которую будут добавляться данные из файлов.
В цикле проходим по каждому файлу в списке.
Загружаем книгу Excel с использованием openpyxl.load_workbook.
Получаем активный лист (первый лист) с использованием loaded_workbook.active.
Используя iter_rows(values_only=True), итерируемся по строкам листа, добавляя значения строк в матрицу data_matrix.
Функция возвращает матрицу с данными из всех файлов.

Пример использования:
'''
files_to_read = ['file1.xlsx', 'file2.xlsx', 'file3.xlsx']
data_matrix = read_excel_files(files_to_read)

'''
Шаг №2: Определение - Сортировка полученной матрицы в порядке убывания
'''
def sort_matrix_descending(matrix):
    # Используем sorted для сортировки по первому элементу в каждой строке в порядке убывания
    sorted_matrix = sorted(matrix, key=lambda x: x[0], reverse=True)
    return sorted_matrix
'''
Описание:

Функция sort_matrix_descending принимает матрицу matrix.
Сортирует матрицу по первому элементу в каждой строке в порядке убывания с использованием функции sorted и 
лямбда-функции lambda x: x[0]. Возвращает отсортированную матрицу sorted_matrix.

Пример использования:
'''
sorted_matrix = sort_matrix_descending(data_matrix)
'''
Шаг №3: Определение - Запись отсортированных данных в один файл, изменение шрифта и добавление границ.
'''
def write_to_excel(output_file, data_matrix):
    new_workbook = openpyxl.Workbook()
    new_sheet = new_workbook.active

    # Записываем отсортированные данные в новый файл
    for row_num, row_data in enumerate(data_matrix, start=1):
        for col_num, value in enumerate(row_data, start=1):
            new_sheet.cell(row=row_num, column=col_num, value=value)

    # Изменим шрифт и добавим границы
    for row in new_sheet.iter_rows(min_row=1, max_row=len(data_matrix), min_col=1, max_col=len(data_matrix[0])):
        for cell in row:
            cell.font = Font(bold=True)
            cell.border = Border(left=Side(border_style='thin'),
                                 right=Side(border_style='thin'),
                                 top=Side(border_style='thin'),
                                 bottom=Side(border_style='thin'))

    # Сохраняем новый файл
    new_workbook.save(output_file)
'''
Описание:

Функция write_to_excel принимает имя файла вывода output_file и отсортированную матрицу data_matrix.
Создает новую книгу Excel и активный лист.
В цикле проходим по отсортированным данным и записываем их в ячейки нового листа.
Изменяем шрифт на жирный и добавляем границы к каждой ячейке.
Сохраняем новый файл с использованием new_workbook.save.

Пример использования:
'''
output_file = 'result.xlsx'
write_to_excel(output_file, sorted_matrix)





